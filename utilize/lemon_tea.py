# 11/20修改：针对新图片修改圆拟合部分（未完成）
# latest version!

import cv2 as cv
import numpy as np
import openpyxl
import os


class Lemon(object):

    def readImage(self, path):
        '''
        Load the image, save its BGR as well as GRAY form as property.
        :param path: Image path
        :return: None
        '''
        img = cv.imread(path)
        self.img = self.rotate_bound(img, 90)  # 向右旋转90度
        gray = cv.cvtColor(self.img, cv.COLOR_BGR2GRAY)
        self.gray = cv.GaussianBlur(gray, (3, 3), 0)

        self.fitting_param = {'file': path.split('\\')[-1]}

    def circle_fitting(self, mark=False):
        '''
        Find the circle contours, and fit into a line.
        :return: thre_masked(for func:line_fitting)
        '''
        # 二值化
        retval, thre = cv.threshold(self.gray, 20, 255, cv.THRESH_BINARY)

        # 寻找铝箔区域
        y = thre.shape[1] // 3
        line = thre[y, :]
        idx = np.where(line == 255)
        x1, x2 = idx[0][0], idx[0][-1]

        # 清除铝箔外区域
        thre[:, :x1] = 0
        thre[:, x2:] = 0

        # 去除最大轮廓
        image, contours, hierarchy = cv.findContours(thre, cv.RETR_TREE, cv.CHAIN_APPROX_NONE)
        cnt_list = []
        for i in contours:
            cnt = {'contour': i, 'area': cv.contourArea(i)}
            cnt_list.append(cnt)
        cnt_list.sort(key=lambda x: x['area'], reverse=True)
        cnt_list.pop(0)

        # 每个轮廓找质心
        # 判断轮廓个数
        mass_center = []
        pre_area = cnt_list[0]['area']
        holes_counts = 0
        for i in cnt_list:
            if i['area'] >= pre_area / 2:
                holes_counts += 1
                pre_area = i['area']
                if mark == True:
                    cv.drawContours(self.img, i['contour'], -1, (0, 0, 255), 2)
                moments = cv.moments(i['contour'])
                x = moments['m10'] / moments['m00']
                y = moments['m01'] / moments['m00']
                mass_center.append((x, y))
            else:
                break

        mass_center.sort(key=lambda x: x[0], reverse=True)
        return mass_center

    def line_fitting(self, thre_masked):
        '''
        Find the line using HoughLinesP method, return the line param
        :return: line param(angle, intercept)
        '''
        edges = cv.Canny(thre_masked, 255, 255)

        minLineLength = 140
        maxLineGap = 100
        lines_num = 0
        angles = []
        intercepts = []

        lines = cv.HoughLinesP(edges, 1, np.pi / 180, 60, minLineLength=minLineLength, maxLineGap=maxLineGap)
        for line in lines:
            x1, y1, x2, y2 = line[0]
            angle = np.arctan2((y2 - y1), (x2 - x1)) * 180.0 / np.pi
            # 双重判断条件
            # 线条位置 & 角度限制
            if x1 <= (self.img.shape[1] / 2) or x2 <= (self.img.shape[1] / 2):
                continue
            if angle >= -5 and angle <= 5:
                cv.line(self.img, (x1, y1), (x2, y2), (0, 0, 255), 2)
                intercept = y2 - (y2 - y1) / (x2 - x1) * x2  # 避免垂直情况出现
                angles.append(angle)
                intercepts.append(intercept)
                lines_num += 1

        # 重绘拟合直线
        rows, cols = self.img.shape[:2]
        avg_angle = np.mean(angles)
        avg_intercept = np.mean(intercepts)
        avg_slope = np.tan(avg_angle / 180.0 * np.pi)
        lefty = int(avg_intercept)
        righty = int(avg_slope * (cols - 1) + avg_intercept)
        cv.line(self.img, (0, lefty), (cols - 1, righty), (255, 0, 0), 2)

        # 在纯黑图上画线
        l_line = np.zeros_like(thre_masked)
        cv.line(l_line, (cols - 1, righty), (0, lefty), 255, 2)

        # # 打印检测统计
        # avg_angle = np.mean(angles)
        # print(file + ':   ', lines_num)
        # print('角度：%s' % angles)
        # print('平均角度：%s' % avg_angle)

        self.fitting_param['wire'] = (avg_angle, avg_intercept)
        self.fitting_param['l_line'] = (lefty, righty)
        return l_line

    def rotate_bound(self, image, angle):
        # grab the dimensions of the image and then determine the
        # center
        (h, w) = image.shape[:2]
        (cX, cY) = (w // 2, h // 2)

        # grab the rotation matrix (applying the negative of the
        # angle to rotate clockwise), then grab the sine and cosine
        # (i.e., the rotation components of the matrix)
        M = cv.getRotationMatrix2D((cX, cY), -angle, 1.0)
        cos = np.abs(M[0, 0])
        sin = np.abs(M[0, 1])

        # compute the new bounding dimensions of the image
        nW = int((h * sin) + (w * cos))
        nH = int((h * cos) + (w * sin))

        # adjust the rotation matrix to take into account translation
        M[0, 2] += (nW / 2) - cX
        M[1, 2] += (nH / 2) - cY

        # perform the actual rotation and return the image
        return cv.warpAffine(image, M, (nW, nH))

    def checkIntersection(self, c_line, l_line):
        # 判断两条线是否相交
        piled_line = cv.bitwise_and(c_line, l_line)
        if 255 in piled_line:
            test.fitting_param['distance'] = 0  # 相交
        else:
            # 求截距
            c_lefty = test.fitting_param['c_line'][0]
            c_righty = test.fitting_param['c_line'][1]
            l_lefty = test.fitting_param['l_line'][0]
            l_righty = test.fitting_param['l_line'][1]
            test.fitting_param['lines_dist'] = np.minimum(np.abs(c_lefty - l_lefty), np.abs(c_righty - l_righty))

    def holesDistance(self, pts):
        cols = self.img.shape[1]
        lefty, righty = self.fitting_param['l_line']
        pts.sort(key=lambda x: x[0], reverse=True)
        distances = []
        for i in pts:
            x, y = i[0], i[1]
            array_longi = np.array([cols - 1, righty - lefty])
            array_trans = np.array([cols - 1 - x, righty - y])
            array_temp = (float(array_trans.dot(array_longi)) / array_longi.dot(array_longi))  # 注意转成浮点数运算
            array_temp = array_longi.dot(array_temp)
            distance = np.sqrt((array_trans - array_temp).dot(array_trans - array_temp))
            distances.append(distance)
        self.fitting_param['holes_dist'] = distances

    def pedalCheck(self, pts):
        pass


if __name__ == '__main__':
    path = 'H:\\bin\\Pieces\\images\\gray\\NG'
    new_path = 'H:\\bin\\Pieces\\images\\gray'
    if not os.path.exists(new_path):
        os.makedirs(new_path)
    dir = os.listdir(path)
    img_list = []
    for i in dir:
        if os.path.splitext(i)[1] == '.bmp':
            img_list.append(i)
    file_num = len(img_list)
    t = 1

    test = Lemon()
    # 新建excel用于保存结果
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    row = 1
    # ws.cell(row=row, column=1).value = '文件名'
    # ws.cell(row=row, column=2).value = '圆夹角'
    # ws.cell(row=row, column=3).value = '圆截距'
    # ws.cell(row=row, column=4).value = '线夹角'
    # ws.cell(row=row, column=5).value = '线截距'
    # ws.cell(row=row, column=6).value = '穿孔数目'
    # ws.cell(row=row, column=7).value = '两线距离'
    # ws.cell(row=row, column=8).value = '左交点（圆）'
    # ws.cell(row=row, column=9).value = '左交点（线）'
    # ws.cell(row=row, column=10).value = '右交点（圆）'
    # ws.cell(row=row, column=11).value = '右交点（线）'
    # ws.cell(row=row, column=12).value = '截距差（圆-线）'

    # 1108
    ws.cell(row=row, column=1).value = '文件名'
    ws.cell(row=row, column=2).value = '距离'

    row += 1
    for file in img_list:
        print(file)
        test.readImage(os.path.join(path, file))
        thre_masked, c_line, pts = test.circle_fitting()
        l_line = test.line_fitting(thre_masked)
        test.checkIntersection(c_line, l_line)
        test.holesDistance(pts)

        # excel保存结果
        # ws.cell(row=row, column=1).value = test.fitting_param['file']
        # ws.cell(row=row, column=2).value = test.fitting_param['holes'][0]
        # ws.cell(row=row, column=3).value = test.fitting_param['holes'][1]
        # ws.cell(row=row, column=4).value = test.fitting_param['wire'][0]
        # ws.cell(row=row, column=5).value = test.fitting_param['wire'][1]
        # ws.cell(row=row, column=6).value = test.fitting_param['holes_num']
        # ws.cell(row=row, column=7).value = test.fitting_param['distance']
        # ws.cell(row=row, column=8).value = test.fitting_param['c_line'][0]
        # ws.cell(row=row, column=9).value = test.fitting_param['l_line'][0]
        # ws.cell(row=row, column=10).value = test.fitting_param['c_line'][1]
        # ws.cell(row=row, column=11).value = test.fitting_param['l_line'][1]
        # ws.cell(row=row, column=12).value = test.fitting_param['holes'][1] - test.fitting_param['wire'][1]

        # 1108
        ws.cell(row=row, column=1).value = test.fitting_param['file']
        ws.cell(row=row, column=2).value = test.fitting_param['holes_dist'][0]
        ws.cell(row=row, column=3).value = test.fitting_param['holes_dist'][1]
        ws.cell(row=row, column=4).value = test.fitting_param['holes_dist'][2]
        ws.cell(row=row, column=5).value = test.fitting_param['holes_dist'][3]
        ws.cell(row=row, column=6).value = test.fitting_param['holes_dist'][4]
        ws.cell(row=row, column=7).value = test.fitting_param['holes_dist'][5]
        row += 1

        # 保存图片
        # cv.imwrite(os.path.join(new_path, file), test.img)

        # 控制台打印进度

        print('已完成：%s / %s' % (t, file_num))
        t += 1

    wb.save(os.path.join(new_path, 'Report.xlsx'))
